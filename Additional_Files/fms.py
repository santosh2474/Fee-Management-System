import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import webbrowser
import os
import random
from PIL import Image, ImageDraw, ImageFont

# Excel file for fee management
fee_excel_file = "Excel_Sheets/student_fee_status.xlsx"

# Create Fee Excel file if it doesn't exist
def create_fee_excel():
    try:
        wb = load_workbook(fee_excel_file)
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Fee Status'
        sheet.append(["Student ID", "Student Name", "Payment To Semester", "Pay Fee", "Due Fee", "Total Fee", "Payment Done By", "Payment Date and Time","Receipt No."])
        wb.save(fee_excel_file)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to create or load the fee status file: {str(e)}")

# Fetch student details by ID from student_registration.xlsx
def fetch_student_details(student_id):
    try:
        wb = load_workbook("Excel_Sheets/student_registration.xlsx")
        sheet = wb['Students']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == student_id:
                return row[1:]  # Return all details except photo and semester
    except FileNotFoundError:
        messagebox.showerror("Error", "Student Registration file not found!")
    return None  # Return None if not found

# Clear all fee fields
def clear_fee_fields():
    entry_student_id.delete(0, tk.END)
    entry_pay_fee.delete(0, tk.END)
    entry_semester_payment.delete(0, tk.END)
    entry_payment_done_by.delete(0, tk.END)  # Clear the payment done by field
    label_name_value.config(text="")
    label_father_value.config(text="")
    label_guardian_contact_value.config(text="")
    label_student_no_value.config(text="")
    label_address_value.config(text="")
    label_course_value.config(text="")
    label_due_fee_value.config(text="")

# Load student details and show in the GUI
def load_student_details():
    student_id = entry_student_id.get()
    details = fetch_student_details(student_id)
    
    if details:
        label_name_value.config(text=details[0])  # Name
        label_father_value.config(text=details[1])  # Father's Name
        label_course_value.config(text=details[2])  # Course

        try:
            total_fee = float(details[4])  # Get total fee
            
            # Calculate the current due fee based on past payments
            wb_fee = load_workbook(fee_excel_file)
            sheet_fee = wb_fee['Fee Status']
            total_paid_fee = sum(float(row[3]) for row in sheet_fee.iter_rows(min_row=2, values_only=True) if row[0] == student_id)
            current_due_fee = total_fee - total_paid_fee

            # Display values in the GUI
            label_due_fee_value.config(text=str(current_due_fee))
            label_student_no_value.config(text=details[5])  # Student Contact No.
            label_guardian_contact_value.config(text=details[6])  # Guardian Contact
            label_address_value.config(text=details[7])  # Address

        except Exception as e:
            messagebox.showerror("Error", f"Error calculating due fee: {str(e)}")
    else:
        messagebox.showerror("Error", "Student ID not found!")


# Function to display the payment receipt with a unique receipt number
def show_payment_receipt(student_id, name, pay_fee, due_fee, total_fee, payment_done_by, payment_date, semester_payment, receipt_no):
    # Create a new window for the receipt
    receipt_window = tk.Toplevel(root)
    receipt_window.title(f"Payment Receipt - Receipt No: {receipt_no}")
    receipt_window.geometry("400x400")  # Set to normal size

    # Make the window always on top
    receipt_window.attributes("-topmost", True)

    # Custom styles
    receipt_label_font = ('Arial', 12, 'bold')
    receipt_entry_font = ('Arial', 12)

    # Receipt details
    tk.Label(receipt_window, text=f"Receipt No: {receipt_no}", font=receipt_label_font).pack(pady=10)
    tk.Label(receipt_window, text=f"Student ID: {student_id}", font=receipt_entry_font).pack(pady=5)
    tk.Label(receipt_window, text=f"Name: {name}", font=receipt_entry_font).pack(pady=5)
    tk.Label(receipt_window, text=f"Payment To Semester: {semester_payment}", font=receipt_entry_font).pack(pady=5)
    tk.Label(receipt_window, text=f"Paid Fee: {pay_fee}", font=receipt_entry_font).pack(pady=5)
    tk.Label(receipt_window, text=f"Payment Done By: {payment_done_by}", font=receipt_entry_font).pack(pady=5)
    tk.Label(receipt_window, text=f"Payment Date: {payment_date}", font=receipt_entry_font).pack(pady=5)

    # Function to print the receipt and save it as an image
    def print_receipt():
        # Create "receipts" folder if it doesn't exist
        receipts_folder = "receipts"
        if not os.path.exists(receipts_folder):
            os.makedirs(receipts_folder)

        # Load logo
        logo_folder = "logo_folder"
        logo_path = os.path.join(logo_folder, "logo.jpg")
        logo = Image.open(logo_path) if os.path.exists(logo_path) else None

        # Create an image for printing (A4 size: 595 x 842 pixels at 72 dpi)
        img = Image.new('RGB', (595, 842), 'white')
        draw = ImageDraw.Draw(img)

        # Draw logo at the top if available
        y_position = 20
        if logo:
            logo.thumbnail((100, 100))
            img.paste(logo, (247, y_position))  # Center the logo at the top
            y_position += 120

        # Add heading below the logo
        heading_text = "Shree Chautara Secondary School, Dandakharka,\n              Sunkoshi-05, Okhaldhunga"
        heading_font = ImageFont.truetype("arial.ttf", 20)
        bbox = draw.textbbox((0, 0), heading_text, font=heading_font)
        w, h = bbox[2] - bbox[0], bbox[3] - bbox[1]
        draw.text(((595 - w) / 2, y_position), heading_text, font=heading_font, fill='black')

        # Add receipt details below the heading
        y_text = y_position + 70
        font_label = ImageFont.truetype("arial.ttf", 14)

        for line in [
            f"Receipt No: {receipt_no}",
            f"Student ID: {student_id}",
            f"Name: {name}",
            f"Payment To Semester: {semester_payment}",
            f"Paid Fee: {pay_fee}",
            f"Payment Done By: {payment_done_by}",
            f"Payment Date: {payment_date}"
        ]:
            draw.text((50, y_text), line, font=font_label, fill='black')
            y_text += 30

        # Draw space for signature
        signature_y = 720  # Position near the bottom of A4 size image
        draw.line((50, signature_y, 545, signature_y), fill='black', width=1)  # Signature line
        draw.text((50, signature_y + 10), "Signature", font=font_label, fill='black')

        # Save the image with the receipt number as the filename
        receipt_file_path = os.path.join(receipts_folder, f"{name}_{receipt_no}.png")
        img.save(receipt_file_path)

        # Print the image (this will open the default image viewer)
        img.show()

    # Add a print button
    print_button = tk.Button(receipt_window, text="Save & Print Receipt", command=print_receipt)
    print_button.pack(pady=20)

# Add fee record to the Excel file and display a message showing paid and due amounts
def add_fee_record():
    student_id = entry_student_id.get()
    pay_fee = entry_pay_fee.get()
    semester_payment = entry_semester_payment.get()
    payment_done_by = entry_payment_done_by.get()  # Get payment done by

    if not student_id or not pay_fee or not semester_payment or not payment_done_by:
        messagebox.showerror("Error", "Please enter Student ID, Pay Fee, Payment To Semester, and Payment Done By!")
        return

    try:
        pay_fee = float(pay_fee)  # Convert to float for calculations
        semester_payment = float(semester_payment)  # Convert semester payment to float
        wb_fee = load_workbook(fee_excel_file)
        sheet_fee = wb_fee['Fee Status']
        payment_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Fetch the total fee from the student registration file
        student_details = fetch_student_details(student_id)
        if student_details:
            total_fee = float(student_details[4])  # The original total fee from student_registration.xlsx

        # Generate a random receipt number
        receipt_no = random.randint(100000, 999999)

        # Calculate new due fee
        current_due_fee = float(label_due_fee_value.cget("text"))
        due_fee = current_due_fee - pay_fee

        # Append the new record with Payment Done By, Payment Date and Time, and Receipt No
        sheet_fee.append([student_id, label_name_value.cget("text"), semester_payment, 
                          pay_fee, due_fee, total_fee, payment_done_by, payment_date, receipt_no])
        wb_fee.save(fee_excel_file)

        # Show payment receipt with all details
        show_payment_receipt(student_id, label_name_value.cget("text"), pay_fee, due_fee, total_fee, payment_done_by, payment_date, semester_payment, receipt_no)

        # Show a message box displaying paid amount and updated due amount
        messagebox.showinfo("Success", f"Fee record added successfully!\nPaid Amount: {pay_fee}\nDue Amount: {due_fee}")

        clear_fee_fields()  # Optionally clear fields after adding
    except ValueError:
        messagebox.showerror("Error", "Invalid fee amount!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to add fee record: {str(e)}")


# Generate HTML report of payment records and save in "Payment Reports" folder
def generate_html_report(student_id):
    try:
        # Create "Payment Reports" folder if it doesn't exist
        report_folder = "Payment Reports"
        if not os.path.exists(report_folder):
            os.makedirs(report_folder)

        wb_fee = load_workbook(fee_excel_file)
        sheet_fee = wb_fee['Fee Status']
        records = [row for row in sheet_fee.iter_rows(min_row=2, values_only=True) if row[0] == student_id]

        if not records:
            messagebox.showinfo("No Records", "No payment records found for this Student ID.")
            return

        # Create HTML content
        html_content = """
        <html>
        <head>
            <title>Payment Records for Student ID {}</title>
            <style>
                table {{ width: 100%; border-collapse: collapse; }}
                th, td {{ border: 1px solid black; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
            </style>
        </head>
        <body>
            <h2>Payment Records for Student ID: {}</h2>
            <table>
                <tr>
                    <th>Student ID</th>
                    <th>Name</th>
                    <th>Payment To Semester</th>
                    <th>Paid Fee</th>
                    <th>Due Fee</th>
                    <th>Total Fee</th>
                    <th>Payment Done By</th>
                    <th>Payment Date and Time</th>
                    <th>Receipt No.</th>
                </tr>
        """.format(student_id, student_id)

        for record in records:
            html_content += "<tr>"
            for item in record:
                html_content += "<td>{}</td>".format(item if item is not None else "")
            html_content += "</tr>"

        html_content += """
            </table>
        </body>
        </html>
        """

        # Save HTML to the "Payment Reports" folder
        html_file = os.path.join(report_folder, "payment_records_{}.html".format(student_id))
        with open(html_file, "w") as f:
            f.write(html_content)

        # Open the HTML file in the default web browser
        webbrowser.open('file://' + os.path.realpath(html_file))

    except Exception as e:
        messagebox.showerror("Error", f"Failed to generate HTML report: {str(e)}")

# Function to open sreg.py
def open_registration():
    try:
        # Close the current main window (main.py)
        root.destroy()
        os.system("python Additional_Files/sreg.py")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Function to open verify.py
def open_verify_receipt():
    try:
        # Close the current main window (main.py)
        root.destroy()
        os.system("python Additional_Files/verify.py")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Function to open today_total.py
def open_today_total():
    try:
        # Close the current main window (main.py)
        root.destroy()
        os.system("python Additional_Files/today_total.py")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Function to open students_details.py
def open_students_details():
    try:
        # Close the current main window (main.py)
        root.destroy()
        os.system("python Additional_Files/students_details.py")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# GUI for Fee Management
root = tk.Tk()
root.title("Student Fee Management")
root.configure(bg='#f0f4f7')

# Set the window to maximized state
root.state('zoomed')

# Custom styles
label_font = ('Arial', 12, 'bold')
entry_font = ('Arial', 12)
button_font = ('Arial', 12, 'bold')

# Main Frame
main_frame = tk.Frame(root, bg='#f0f4f7')
main_frame.pack(pady=20)

# Fee Management form
tk.Label(main_frame, text="Student ID", font=label_font).grid(row=0, column=0, padx=10, pady=5, sticky='e')
entry_student_id = tk.Entry(main_frame, font=entry_font)
entry_student_id.grid(row=0, column=1, padx=10, pady=5)

tk.Button(main_frame, text="Load Details", font=button_font, command=load_student_details).grid(row=0, column=2, padx=10, pady=5)

tk.Label(main_frame, text="Name:", font=label_font).grid(row=1, column=0, padx=10, pady=5, sticky='e')
label_name_value = tk.Label(main_frame, text="", font=entry_font)
label_name_value.grid(row=1, column=1, padx=10, pady=5)

tk.Label(main_frame, text="Father's Name:", font=label_font).grid(row=2, column=0, padx=10, pady=5, sticky='e')
label_father_value = tk.Label(main_frame, text="", font=entry_font)
label_father_value.grid(row=2, column=1, padx=10, pady=5)

tk.Label(main_frame, text="Guardian Contact:", font=label_font).grid(row=3, column=0, padx=10, pady=5, sticky='e')
label_guardian_contact_value = tk.Label(main_frame, text="", font=entry_font)
label_guardian_contact_value.grid(row=3, column=1, padx=10, pady=5)

tk.Label(main_frame, text="Student No:", font=label_font).grid(row=4, column=0, padx=10, pady=5, sticky='e')
label_student_no_value = tk.Label(main_frame, text="", font=entry_font)
label_student_no_value.grid(row=4, column=1, padx=10, pady=5)

tk.Label(main_frame, text="Address:", font=label_font).grid(row=5, column=0, padx=10, pady=5, sticky='e')
label_address_value = tk.Label(main_frame, text="", font=entry_font)
label_address_value.grid(row=5, column=1, padx=10, pady=5)

tk.Label(main_frame, text="Course:", font=label_font).grid(row=6, column=0, padx=10, pady=5, sticky='e')
label_course_value = tk.Label(main_frame, text="", font=entry_font)
label_course_value.grid(row=6, column=1, padx=10, pady=5)

tk.Label(main_frame, text="Due Fee:", font=label_font).grid(row=7, column=0, padx=10, pady=5, sticky='e')
label_due_fee_value = tk.Label(main_frame, text="", font=entry_font)
label_due_fee_value.grid(row=7, column=1, padx=10, pady=5)

tk.Label(main_frame, text="Pay Fee:", font=label_font).grid(row=8, column=0, padx=10, pady=5, sticky='e')
entry_pay_fee = tk.Entry(main_frame, font=entry_font)
entry_pay_fee.grid(row=8, column=1, padx=10, pady=5)

tk.Label(main_frame, text="Payment To Semester:", font=label_font).grid(row=9, column=0, padx=10, pady=5, sticky='e')
entry_semester_payment = tk.Entry(main_frame, font=entry_font)
entry_semester_payment.grid(row=9, column=1, padx=10, pady=5)

tk.Label(main_frame, text="Payment Done By:", font=label_font).grid(row=10, column=0, padx=10, pady=5, sticky='e')
entry_payment_done_by = tk.Entry(main_frame, font=entry_font)
entry_payment_done_by.grid(row=10, column=1, padx=10, pady=5)


# Buttons
button_frame = tk.Frame(main_frame, bg='#f0f4f7')
button_frame.grid(row=11, column=0, columnspan=3, pady=20)
# Existing buttons
tk.Button(button_frame, text="Add Fee Record", font=button_font, command=add_fee_record).grid(row=0, column=1, padx=40)
tk.Button(button_frame, text="Evaluate Daily Payments", font=button_font, command=open_today_total).grid(row=1, column=0, padx=20, pady=20)
tk.Button(button_frame, text="Generate Payment Report", font=button_font, command=lambda: generate_html_report(entry_student_id.get())).grid(row=1, column=2, padx=20, pady=20)

# New buttons for Registration and Verify
tk.Button(button_frame, text="Registration", font=button_font, command=open_registration).grid(row=2, column=0, padx=20, pady=(10, 0))
tk.Button(button_frame, text="Registred Student Details", font=button_font, command=open_students_details).grid(row=1, column=1, padx=20, pady=(1, 0))
tk.Button(button_frame, text="Verify Receipt", font=button_font, command=open_verify_receipt).grid(row=2, column=2, padx=20, pady=(10, 0))

# Create Excel file if not exists
create_fee_excel()

# Function to handle Exit
def on_closing():
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        root.destroy()

root.protocol("WM_DELETE_WINDOW", on_closing) 

root.mainloop()
