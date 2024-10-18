import sys
import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk
import os
from datetime import datetime


# File paths
excel_file = "Excel_Sheets/student_registration.xlsx"
photos_directory = "student_photos"

# Create Excel file and directory if they do not exist
def create_excel():
    try:
        wb = load_workbook(excel_file)
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Students'
        sheet.append(["Student ID", "Name", "Father's Name", "Course", "Semester", "Total Fees", "Student Contact No.",
              "Guardian Contact No.", "Address", "Photo Path", "Registration Date", "Date of Birth (BS) [YYYY-MM-DD]"])
        wb.save(excel_file)

    if not os.path.exists(photos_directory):
        os.makedirs(photos_directory)

create_excel()


def clear_fields():
    """Clears all input fields after adding or updating a record"""
    entry_student_id.delete(0, tk.END)
    entry_name.delete(0, tk.END)
    entry_father_name.delete(0, tk.END)
    entry_course.delete(0, tk.END)
    entry_semester.delete(0, tk.END)
    entry_total_fees.delete(0, tk.END)
    entry_student_contact.delete(0, tk.END)
    entry_guardian_contact.delete(0, tk.END)
    entry_address.delete(0, tk.END)
    entry_dob.delete(0, tk.END)
    label_photo.config(image='')
    global photo_path
    photo_path = None

def upload_photo():
    global photo_path
    student_name = entry_name.get().strip()
    student_id = entry_student_id.get().strip()
    
    if not student_name or not student_id:
        messagebox.showerror("Error", "Please enter the student's name and ID before uploading a photo!")
        return
    
    file_path = filedialog.askopenfilename(title="Select Photo", filetypes=[("Image Files", "*.jpg *.jpeg *.png")])
    if file_path:
        img = Image.open(file_path)
        
        # Convert image to 'RGB' if it has an alpha channel (RGBA)
        if img.mode == 'RGBA':
            img = img.convert('RGB')
        
        img.thumbnail((150, 150))
        photo_img = ImageTk.PhotoImage(img)
        label_photo.config(image=photo_img)
        label_photo.image = photo_img

        # Create a unique filename using the student's name and ID
        photo_name = f"{student_name.replace(' ', '_')}_{student_id}.jpg"
        photo_path = os.path.join(photos_directory, photo_name)
        img.save(photo_path)
        
        messagebox.showinfo("Photo Uploaded", "Photo uploaded successfully!")

def add_record():
    student_id = entry_student_id.get().strip()
    name = entry_name.get().strip()
    father_name = entry_father_name.get().strip()
    course = entry_course.get().strip()
    semester = entry_semester.get().strip()
    total_fees = entry_total_fees.get().strip()
    student_contact = entry_student_contact.get().strip()
    guardian_contact = entry_guardian_contact.get().strip()
    address = entry_address.get().strip()
    registration_date = datetime.now().strftime("%Y-%m-%d")  # Set the current date
    dob = entry_dob.get().strip()  # Get the DOB from the input field

    if not all([student_id, name, father_name, course, semester, total_fees, student_contact, guardian_contact, address, photo_path, registration_date, dob]):
        messagebox.showerror("Error", "Please fill in all fields, including the registration date, date of birth, and upload a photo!")
        return

    new_data = [student_id, name, father_name, course, semester, total_fees, student_contact, guardian_contact, address, photo_path, registration_date, dob]

    try:
        wb = load_workbook(excel_file)
        sheet = wb['Students']

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == student_id:
                messagebox.showerror("Error", "Student ID already exists!")
                return

        sheet.append(new_data)
        wb.save(excel_file)
        
        label_registration_date.config(text=f"Registration Date: {registration_date}")  # Update the label with the current date
        messagebox.showinfo("Success", "Record Added Successfully!")
        clear_fields()
    except Exception as e:
        messagebox.showerror("Error", str(e))

def search_record():
    student_id = entry_student_id.get().strip()
    if not student_id:
        messagebox.showerror("Error", "Please enter Student ID to search!")
        return
    
    try:
        wb = load_workbook(excel_file)
        sheet = wb['Students']
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == student_id:
                entry_name.delete(0, tk.END)
                entry_name.insert(0, row[1])
                entry_father_name.delete(0, tk.END)
                entry_father_name.insert(0, row[2])
                entry_course.delete(0, tk.END)
                entry_course.insert(0, row[3])
                entry_semester.delete(0, tk.END)
                entry_semester.insert(0, row[4])
                entry_total_fees.delete(0, tk.END)
                entry_total_fees.insert(0, row[5])
                entry_student_contact.delete(0, tk.END)
                entry_student_contact.insert(0, row[6])
                entry_guardian_contact.delete(0, tk.END)
                entry_guardian_contact.insert(0, row[7])
                entry_address.delete(0, tk.END)
                entry_address.insert(0, row[8])
                entry_dob.delete(0, tk.END)
                entry_dob.insert(0, row[11])  # Assuming DOB is in the 12th column

                global photo_path
                photo_path = row[9]
                if photo_path and os.path.exists(photo_path):
                    img = Image.open(photo_path)
                    img.thumbnail((150, 150))
                    photo_img = ImageTk.PhotoImage(img)
                    label_photo.config(image=photo_img)
                    label_photo.image = photo_img
                
                label_registration_date.config(text=f"Registration Date: {row[10]}")  # Update the label's text

                return
        messagebox.showinfo("Not Found", "No record found with the given Student ID.")
    except Exception as e:
        messagebox.showerror("Error", str(e))

def update_record():
    student_id = entry_student_id.get().strip()
    if not student_id:
        messagebox.showerror("Error", "Please enter Student ID to update!")
        return
    
    try:
        wb = load_workbook(excel_file)
        sheet = wb['Students']
        
        for row in sheet.iter_rows(min_row=2, max_col=10):
            if row[0].value == student_id:
                row[1].value = entry_name.get()
                row[2].value = entry_father_name.get()
                row[3].value = entry_course.get()
                row[4].value = entry_semester.get()
                row[5].value = entry_total_fees.get()
                row[6].value = entry_student_contact.get()
                row[7].value = entry_guardian_contact.get()
                row[8].value = entry_address.get()
                row[9].value = photo_path
                row[10].value = label_registration_date.get()
                row[11].value = entry_dob.get()
                wb.save(excel_file)
                messagebox.showinfo("Success", "Record Updated Successfully!")
                clear_fields()
                return
        
        messagebox.showerror("Error", "No record found with the given Student ID.")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Function to open fms.py
def open_payment_fee():
    try:
        # Close the current main window (main.py)
        root.destroy()
        os.system("python Additional_Files/fms.py")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Function to open verify.py
def open_verify_receipt():
    try:
        # Close the current main window (main.py)
        root.destroy()
        os.system("python Additional_Files/verify.py")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Function to open students_details.py
def open_students_details():
    try:
        root.destroy()
        os.system("python Additional_Files/students_details.py")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Password Authentication
def authenticate():
    if password_entry.get() == "admin123":
        password_window.destroy()
        main_window()
    else:
        messagebox.showerror("Error", "Incorrect password!")

def password_prompt():
    global password_entry, password_window

    # Create the main password window
    password_window = tk.Tk()
    password_window.title("Authentication")
    password_window.state('zoomed')  # Full-screen window
    password_window.configure(bg='#f0f0f5')  # Light grey background

    # Create a frame in the center to hold the elements
    frame = tk.Frame(password_window, bg='#e8eaf6', padx=50, pady=50)
    frame.place(relx=0.5, rely=0.5, anchor='center')  # Center the frame

    # Title Label
    tk.Label(
        frame, text="Student Registration System", font=("Arial", 28, "bold"), 
        bg='#e8eaf6', fg='#1a237e'
    ).pack(pady=(0, 20))  # Extra padding between title and input field

    # Instruction Label
    tk.Label(
        frame, text="Enter Password:", font=("Arial", 18), 
        bg='#e8eaf6', fg='#3e2723'
    ).pack(pady=10)

    # Password Entry Box
    password_entry = tk.Entry(frame, show="*", font=("Arial", 16), width=25)
    password_entry.pack(pady=10)

    # Login Button
    tk.Button(
        frame, text="Login", font=("Arial", 16, "bold"), width=10,
        bg='#81c784', fg='white', activebackground='#66bb6a',
        command=authenticate
    ).pack(pady=20)

    # Close Button for convenience
    tk.Button(
        frame, text="Exit", font=("Arial", 12), width=8,
        bg='#ef9a9a', fg='white', activebackground='#e57373',
        command=password_window.destroy
    ).pack()

    password_window.mainloop()

# Placeholder for the main application window
def main_window():
    main = tk.Tk()
    main.title("Main Window")
    main.geometry("800x600")
    tk.Label(main, text="Welcome to the Student Registration System", font=("Arial", 24)).pack(pady=50)
    main.mainloop()

# GUI Window
def main_window():
    global root, entry_student_id, entry_name, entry_father_name, entry_course, entry_semester
    global entry_total_fees, entry_student_contact, entry_guardian_contact, entry_address, entry_dob, label_photo, photo_path, label_registration_date

    root = tk.Tk()
    root.title("Student Registration")
    root.state('zoomed')  # Make the window fullscreen
    root.configure(bg='#e0f7fa')

    # Function to handle Exit
    def on_closing():
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_closing) 

    # Main Frame
    main_frame = tk.Frame(root, bg='#e0f7fa')
    main_frame.pack(pady=20)

    # Registration Form Labels and Entries
    label_font = ('Arial', 12, 'bold')
    entry_font = ('Arial', 12)
    button_font = ('Arial', 12, 'bold')

    fields = ["Student ID", "Name", "Father's Name", "Course", "Semester", "Total Fees", "Student Contact No.", "Guardian Contact No.", "Address", "Date of Birth (BS) [YYYY-MM-DD]"]
    entries = {}

    for i, field in enumerate(fields):
        tk.Label(main_frame, text=field, font=label_font, bg='#e0f7fa').grid(row=i, column=0, padx=10, pady=5, sticky='e')
        entry = tk.Entry(main_frame, font=entry_font)
        entry.grid(row=i, column=1, padx=10, pady=5, sticky='w')
        entries[field] = entry

    entry_student_id = entries["Student ID"]
    entry_name = entries["Name"]
    entry_father_name = entries["Father's Name"]
    entry_course = entries["Course"]
    entry_semester = entries["Semester"]
    entry_total_fees = entries["Total Fees"]
    entry_student_contact = entries["Student Contact No."]
    entry_guardian_contact = entries["Guardian Contact No."]
    entry_address = entries["Address"]
    entry_dob = entries["Date of Birth (BS) [YYYY-MM-DD]"]  # New DOB Entry

    # Upload Photo and Display Area
    label_photo = tk.Label(main_frame, bg='#e0f7fa')
    label_photo.grid(row=len(fields), column=0, columnspan=3, pady=5)

    # Upload Photo button in the middle
    tk.Button(main_frame, text="Upload Photo", font=button_font, command=upload_photo).grid(row=len(fields)+1, column=0, columnspan=3, pady=5)

    # Display the registration date below the photo area
    label_registration_date = tk.Label(main_frame, text="Registration Date: ", font=label_font, bg='#e0f7fa')
    label_registration_date.grid(row=len(fields)+2, column=0, columnspan=3, pady=10)

    # Button Frame for action buttons
    button_frame = tk.Frame(root, bg='#e0f7fa')
    button_frame.pack(pady=10)
    
    buttons_top = [
        ("Add", add_record),
        ("Search", search_record),
        ("Update", update_record)
    ]

    for text, command in buttons_top:
        tk.Button(button_frame, text=text, font=button_font, command=command, width=15).pack(side='left', padx=10)

    # Button Frame for Payment Fee and Verify Receipt (in a separate row)
    button_frame_bottom = tk.Frame(root, bg='#e0f7fa')
    button_frame_bottom.pack(pady=10)

    buttons_bottom = [
        ("Payment Fee", open_payment_fee),
        ("Verify Receipt", open_verify_receipt),
        ("Registred Students Detail", open_students_details)
    ]

    for text, command in buttons_bottom:
        tk.Button(button_frame_bottom, text=text, font=button_font, command=command, width=20).pack(side='left', padx=10)

    root.mainloop()


password_prompt()