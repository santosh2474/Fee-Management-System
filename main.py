import sys
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Sample user credentials (you can replace this with a database or file system in a real application)
user_credentials = {
    "admin": "password",
    "user1": "pass1",
    "user2": "pass2",
}

# Path to Excel log file
excel_file_path = os.path.join("Excel_Sheets", "user_logs.xlsx")

# Create a new Excel file if it doesn't exist
if not os.path.exists(excel_file_path):
    os.makedirs("Excel_Sheets", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.append(["User ID", "Login Time"])  # Add header row
    wb.save(excel_file_path)

# Function to log user activity
def log_user_activity(user_id, login_time):
    wb = load_workbook(excel_file_path)
    ws = wb.active
    ws.append([user_id, login_time])  # No logout time recorded
    wb.save(excel_file_path)

# Create the main window
root = tk.Tk()
root.title("Fees Management System")

# Start the window in fullscreen mode
root.state('zoomed')  # Start maximized

# Create a frame to contain the widgets
main_frame = tk.Frame(root, bg="#f0f0f0")
main_frame.pack(expand=True, fill="both")

# Load and place the logo
def load_logo(size_factor=0.15):
    logo_path = os.path.join("logo_folder", "logo.jpg")  # Path to the logo file
    try:
        logo_image = Image.open(logo_path)
        screen_width = root.winfo_screenwidth()  # Get screen width for scaling
        size = int(screen_width * size_factor)  # Adjust size based on screen width
        logo_image = logo_image.resize((size, size), Image.Resampling.LANCZOS)  # Make the logo square
        logo = ImageTk.PhotoImage(logo_image)
        return logo  # Return logo for later use
    except FileNotFoundError:
        messagebox.showerror("Error", "Logo image not found!")
        return None

# Function to add heading above the logo
def add_heading():
    large_text = "श्री चौतारा माध्यमिक विद्यालय डाँडाखर्क सुनकोशी -५, ओखलढुंगा"
    small_text = ("दक्ष जनशक्ति र प्रबिधिको उपयोगको लागि ऐक्यबद्धता\n"
                  "गुणस्तरीय र सर्वसुलभ शिक्षा हाम्रो प्रतिबद्धता")

    large_text_label = tk.Label(main_frame, text=large_text, bg="#f0f0f0", fg="blue", font=("Arial", 20, "bold"), justify="center")
    large_text_label.pack(pady=(10, 0), padx=20, anchor=tk.CENTER)
    
    small_text_label = tk.Label(main_frame, text=small_text, bg="#f0f0f0", fg="#7098AB", font=("Arial", 16, "bold"), justify="center")
    small_text_label.pack(pady=(0, 10), padx=20, anchor=tk.CENTER)

# Function to toggle password visibility
def toggle_password_visibility():
    if password_entry.cget('show') == '*':
        password_entry.config(show='')  
        toggle_button.config(text="👁")  
    else:
        password_entry.config(show='*')  
        toggle_button.config(text="👁‍🗨")  

# Function to display logo and heading in the login window
def show_login_window():
    global password_entry, toggle_button, current_user_id  # Declare global variables
    for widget in main_frame.winfo_children():
        widget.destroy()

    add_heading()
    logo = load_logo(size_factor=0.15)
    if logo:
        logo_label = tk.Label(main_frame, image=logo, bg="#f0f0f0")
        logo_label.image = logo  
        logo_label.pack(pady=(0, 5), anchor=tk.CENTER)  

    login_frame = tk.Frame(main_frame, bg="#f0f0f0")
    login_frame.pack(expand=True)

    title_label = tk.Label(login_frame, text="Fee Management System", 
                           bg="#f0f0f0", fg="#3498db", font=("Arial", 24, "bold"))
    title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20), padx=20, sticky="ew")

    # User ID Section
    user_id_label = tk.Label(login_frame, text="👤 User ID:", bg="#f0f0f0", font=("Arial", 18))
    user_id_label.grid(row=1, column=0, pady=10, padx=(10, 0), sticky="e")

    user_id_entry = tk.Entry(login_frame, font=("Arial", 18), width=25)
    user_id_entry.grid(row=1, column=1, pady=10, padx=(0, 10), sticky="w")

    # Password Section
    password_label = tk.Label(login_frame, text="🔒 Password:", bg="#f0f0f0", font=("Arial", 18))
    password_label.grid(row=2, column=0, pady=10, padx=(10, 0), sticky="e")

    password_frame = tk.Frame(login_frame, bg="#f0f0f0")  
    password_frame.grid(row=2, column=1, pady=10, padx=(0, 10), sticky="w")

    password_entry = tk.Entry(password_frame, show='*', font=("Arial", 18), width=22)
    password_entry.pack(side="left", padx=(0, 5))

    toggle_button = tk.Button(password_frame, text="👁‍🗨", 
                              command=toggle_password_visibility, font=("Arial", 15), 
                              bg="#f0f0f0", relief="flat", bd=0)
    toggle_button.pack(side="left")

    # Login Button
    def validate_login():
        user_id = user_id_entry.get()
        password = password_entry.get()
        if user_id in user_credentials and user_credentials[user_id] == password:  
            global current_user_id, login_time
            current_user_id = user_id
            login_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_user_activity(current_user_id, login_time)  # Log login time
            create_main_buttons()  
        else:
            messagebox.showerror("Login Failed", "Invalid User ID or Password")

    login_button = tk.Button(login_frame, text="Login", command=validate_login, 
                             bg="#3498db", fg="white", font=("Arial", 18), width=20, height=2)
    login_button.grid(row=3, column=0, columnspan=2, pady=20)

def create_main_buttons():
    for widget in main_frame.winfo_children():
        widget.destroy()

    add_heading()
    logo = load_logo()
    if logo:
        logo_label = tk.Label(main_frame, image=logo, bg="#f0f0f0")
        logo_label.image = logo  
        logo_label.pack(pady=(10, 10), anchor=tk.CENTER)

    create_buttons()

def open_registration():
    try:
        root.destroy()
        os.execl(sys.executable, sys.executable, os.path.join(os.getcwd(), 'Additional_Files/sreg.py'))
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open registration window: {e}")

def open_fee_payment():
    try:
        root.destroy()
        os.execl(sys.executable, sys.executable, os.path.join(os.getcwd(), 'Additional_Files/fms.py'))
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open fee payment window: {e}")

def open_receipt_verify():
    try:
        root.destroy()
        os.execl(sys.executable, sys.executable, os.path.join(os.getcwd(), 'Additional_Files/verify.py'))
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open receipt verification window: {e}")

# Function to handle Exit
def on_closing():
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        root.destroy()

root.protocol("WM_DELETE_WINDOW", on_closing)  

button_style = {
    "bg": "#3498db",  
    "fg": "white",    
    "font": ("Arial", 20, "bold"),  
    "width": 30,      
    "height": 2,      
    "bd": 0,          
    "activebackground": "#2980b9"  
}

def create_buttons():
    button_frame = tk.Frame(main_frame, bg="#f0f0f0")
    button_frame.pack(pady=20, anchor=tk.CENTER)

    btn_registration = tk.Button(button_frame, text="Registration", **button_style, command=open_registration)
    btn_registration.pack(pady=10)

    btn_fee_payment = tk.Button(button_frame, text="Fee Payment", **button_style, command=open_fee_payment)
    btn_fee_payment.pack(pady=10)

    btn_receipt_verify = tk.Button(button_frame, text="Receipt Verification", **button_style, command=open_receipt_verify)
    btn_receipt_verify.pack(pady=10)


# Show login window
show_login_window()

# Start the Tkinter main loop
root.mainloop()
